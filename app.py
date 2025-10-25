# app.py
import os
from flask import Flask, request, jsonify, render_template, send_from_directory, session
from werkzeug.utils import secure_filename
from firebase_admin import credentials, initialize_app, auth, firestore, storage
import firebase_admin
import pandas as pd
from auth_utils import admin_required, login_required, web_login_required
from flask_cors import CORS
from processing.file_processors import process_uploaded_file  # placeholder
from firebase_config import db, bucket
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import base64
from datetime import datetime
import time
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import json
import os

# Add these configurations after your existing config
GOOGLE_SHEETS_SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# Initialize Google Sheets client


UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = set(['xlsx','xls','csv','png','jpg','jpeg','pdf','txt','mp4','mp3'])

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-this')  # Add this for sessions
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def init_google_sheets():
    try:
        # For Render deployment, use environment variable
        creds_json = os.environ.get('GOOGLE_SHEETS_CREDENTIALS')
        if creds_json:
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=GOOGLE_SHEETS_SCOPES)
        else:
            # For local development, use file
            creds = Credentials.from_service_account_file('test123456.json', scopes=GOOGLE_SHEETS_SCOPES)

        
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        print(f"Error initializing Google Sheets: {e}")
        return None

# Global sheets client
sheets_client = init_google_sheets()
PENDING_SHEET_ID = os.environ.get("PENDING_SHEET_ID")  # Replace with your actual sheet ID

# Add these new routes to your app.py

@app.route('/staff-registration')
def staff_registration():
    return render_template('staff_registration.html')

@app.route('/api/submit_registration', methods=['POST'])
def submit_registration():
    try:
        if not sheets_client:
            return jsonify({'success': False, 'error': 'Google Sheets not configured'})
        
        data = request.get_json()
        
        # Validate required fields - email is now primary identifier
        required_fields = ['name', 'email', 'department', 'mobile_no', 'designation']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field gjhgjhg: {field}'})
        
        # Open the pending registrations sheet
        sheet = sheets_client.open_by_key(PENDING_SHEET_ID).sheet1
        
        # Check if email already exists in pending (use email as unique identifier)
        existing_records = sheet.get_all_records()
        for record in existing_records:
            if record.get('Email') == data.get('email'):
                return jsonify({'success': False, 'error': 'Registration already exists for this email address'})
        
        # Also check if employee number exists if provided
        if data.get('emp_no'):
            for record in existing_records:
                if record.get('Employee ID') == data.get('emp_no'):
                    return jsonify({'success': False, 'error': 'Registration already exists for this Employee ID'})
        
        # Prepare row data (match your sheet columns)
        row_data = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Timestamp
            data.get('name', ''),
            data.get('emp_no', ''),  # Can be empty now
            data.get('email', ''),
            data.get('department', ''),
            data.get('designation', ''),
            data.get('mobile_no', ''),
            data.get('type', ''),
            data.get('contract_type', ''),
            data.get('category', ''),
            data.get('gender', ''),
            data.get('blood_group', ''),
            data.get('permanent_address', ''),
            'PENDING'  # Status column
        ]
        
        # Add row to sheet
        sheet.append_row(row_data)
        
        return jsonify({'success': True, 'message': 'Registration submitted successfully! Please wait for admin approval.'})
        
    except Exception as e:
        print(f"Registration submission error: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/pending_registrations', methods=['GET'])
@admin_required
def get_pending_registrations():
    try:
        if not sheets_client:
            return jsonify({'success': False, 'error': 'Google Sheets not configured'})
        
        # Open the pending registrations sheet
        sheet = sheets_client.open_by_key(PENDING_SHEET_ID).sheet1
        
        # Get all records
        records = sheet.get_all_records()
        
        # Filter only pending records
        pending_records = [record for record in records if record.get('Status') == 'PENDING']
        
        return jsonify({'success': True, 'registrations': pending_records})
        
    except Exception as e:
        print(f"Error fetching pending registrations: {e}")
        return jsonify({'success': False, 'error': str(e)})

        
@app.route('/api/approve_registration', methods=['POST'])
@admin_required
def approve_registration():
    try:
        if not sheets_client:
            return jsonify({'success': False, 'error': 'Google Sheets not configured'})
        
        data = request.get_json()
        # Use email as primary identifier instead of emp_no
        email = data.get('email')
        
        if not email:
            return jsonify({'success': False, 'error': 'Email address required'})
        
        # Open the pending registrations sheet
        sheet = sheets_client.open_by_key(PENDING_SHEET_ID).sheet1
        records = sheet.get_all_records()
        
        # Find the record to approve using email
        record_to_approve = None
        row_index = None
        
        for idx, record in enumerate(records, start=2):  # Start from 2 because row 1 is header
            if record.get('Email') == email and record.get('Status') == 'PENDING':
                record_to_approve = record
                row_index = idx
                break
        
        if not record_to_approve:
            return jsonify({'success': False, 'error': 'Registration not found'})
        
        # Add to main database
        staff_data = {
            'slNo': get_next_sl_no(),
            'empNo': record_to_approve.get('Employee ID', ''),  # Can be empty
            'name': record_to_approve.get('Name', ''),
            'type': record_to_approve.get('Type', ''),
            'contractType': record_to_approve.get('Contract Type', ''),
            'department': record_to_approve.get('Department', ''),
            'category': record_to_approve.get('Category', ''),
            'gender': record_to_approve.get('Gender', ''),
            'designation': record_to_approve.get('Designation', ''),
            'mobileNo': record_to_approve.get('Mobile No', ''),
            'bloodGroup': record_to_approve.get('Blood Group', ''),
            'permanentAddress': record_to_approve.get('Permanent Address', ''),
            'email': record_to_approve.get('Email', ''),
            'photo': ''  # Empty for now
        }
        
        # Use email as doc ID since emp_no might be empty
        doc_id = staff_data['email'].replace('@', '_at_').replace('.', '_dot_')  # Make email Firebase-safe
        doc_ref = db.collection('staff').document(doc_id)
        
        # Check if staff already exists in database
        if doc_ref.get().exists:
            return jsonify({'success': False, 'error': 'Staff member already exists in database'})
        
        # Set the document
        doc_ref.set(staff_data)
        
        # Verify the write succeeded
        if not doc_ref.get().exists:
            raise Exception('Failed to verify staff addition in Firestore')
        
        # Delete the row from Google Sheets
        sheet.delete_rows(row_index)
        
        return jsonify({'success': True, 'message': 'Registration approved and added to database'})
        
    except Exception as e:
        print(f"Error approving registration: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/reject_registration', methods=['POST'])
@admin_required
def reject_registration():
    try:
        if not sheets_client:
            return jsonify({'success': False, 'error': 'Google Sheets not configured'})
        
        data = request.get_json()
        # Use email as primary identifier instead of emp_no
        email = data.get('email')
        
        if not email:
            return jsonify({'success': False, 'error': 'Email address required'})
        
        # Open the pending registrations sheet
        sheet = sheets_client.open_by_key(PENDING_SHEET_ID).sheet1
        records = sheet.get_all_records()
        
        # Find and delete the record using email
        for idx, record in enumerate(records, start=2):  # Start from 2 because row 1 is header
            if record.get('Email') == email and record.get('Status') == 'PENDING':
                sheet.delete_rows(idx)
                return jsonify({'success': True, 'message': 'Registration rejected and removed'})
        
        return jsonify({'success': False, 'error': 'Registration not found'})
        
    except Exception as e:
        print(f"Error rejecting registration: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Helper function to get next serial number
def get_next_sl_no():
    try:
        staffs = db.collection('staff').get()
        if not staffs:
            return 1
        
        max_sl = 0
        for staff in staffs:
            staff_data = staff.to_dict()
            sl_no = staff_data.get('slNo', 0)
            if isinstance(sl_no, (int, str)) and str(sl_no).isdigit():
                max_sl = max(max_sl, int(sl_no))
        
        return max_sl + 1
    except:
        return 1


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('login.html')


# Updated Excel upload endpoint to match Flutter functionality exactly
@app.route('/api/upload_excel', methods=['POST'])
@admin_required
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    
    filename = secure_filename(f.filename)
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(path)

    # Required columns exactly as in Flutter
    required_columns = [
        'Sl No', 'Emp No', 'Name', 'Type', 'Contract Type', 'Department', 
        'Category', 'Gender', 'Designation', 'Mobile No', 'Blood Group', 
        'Permanent Address', 'Email', 'Photo'
    ]

    try:
        # Load workbook to handle images
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.strip() if cell.value else '')
        
        # Create column indices mapping
        column_indices = {}
        for i, header in enumerate(headers):
            if header:
                column_indices[header] = i
        
        # Check for missing required columns (Photo is optional)
        missing_columns = []
        for col in required_columns:
            if col != 'Photo' and col not in column_indices:
                missing_columns.append(col)
        
        if missing_columns:
            return jsonify({
                'error': f'Missing columns: {", ".join(missing_columns)}'
            }), 400
        
        has_photo_column = 'Photo' in column_indices
        staff_data = []
        total_records = sheet.max_row - 1  # Exclude header row
        processed_records = 0
        errors = []
        
        # Process data rows (starting from row 2)
        for row_num in range(2, sheet.max_row + 1):
            try:
                row = sheet[row_num]
                
                # Skip empty rows
                if not row[0].value:
                    total_records -= 1
                    continue
                
                staff = {}
                photo_url = ''
                
                # Process text columns
                for column in required_columns:
                    if column == 'Photo':
                        continue
                    
                    if column not in column_indices:
                        staff[column] = ''
                        continue
                    
                    col_index = column_indices[column]
                    if col_index >= len(row):
                        staff[column] = ''
                        continue
                    
                    cell_value = row[col_index].value
                    
                    # Special handling for Mobile No
                    if column == 'Mobile No' and cell_value:
                        cell_value = str(cell_value).replace('.0', '')
                    
                    staff[column] = str(cell_value).strip() if cell_value else ''
                
                # Handle photo if column exists
                if has_photo_column:
                    photo_col_index = column_indices['Photo']
                    
                    # Check if there's an image in this cell
                    photo_url = process_excel_image(sheet, row_num, photo_col_index, staff.get('Email', staff.get('Sl No', str(int(time.time())))))
                
                staff['photoUrl'] = photo_url
                staff_data.append(staff)
                processed_records += 1
                
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e)
                })
        
        # Batch upload to Firestore exactly like Flutter
        batch = db.batch()
        
        for staff in staff_data:
            # Use Email as document ID if available, otherwise use Emp No, otherwise auto-generate
            doc_id = None
            if staff.get('Email') and staff.get('Email').strip():
                # Make email Firebase-safe by replacing special characters
                doc_id = staff.get('Email').strip().replace('@', '_at_').replace('.', '_dot_')
            elif staff.get('Emp No') and staff.get('Emp No').strip():
                doc_id = str(staff.get('Emp No')).strip()
            
            if doc_id:
                doc_ref = db.collection('staff').document(doc_id)
            else:
                doc_ref = db.collection('staff').document()
            
            # Structure exactly matching Flutter
            batch.set(doc_ref, {
                'slNo': staff.get('Sl No', ''),
                'empNo': staff.get('Emp No', ''),
                'name': staff.get('Name', ''),
                'type': staff.get('Type', ''),
                'contractType': staff.get('Contract Type', ''),
                'department': staff.get('Department', ''),
                'category': staff.get('Category', ''),
                'gender': staff.get('Gender', ''),
                'designation': staff.get('Designation', ''),
                'mobileNo': staff.get('Mobile No', ''),
                'bloodGroup': staff.get('Blood Group', ''),
                'permanentAddress': staff.get('Permanent Address', ''),
                'email': staff.get('Email', ''),
                'photoUrl': staff.get('photoUrl', ''),
                'timestamp': firestore.SERVER_TIMESTAMP,
            })
        
        # Commit batch
        batch.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded {len(staff_data)} staff records!',
            'totalRecords': total_records,
            'processedRecords': processed_records,
            'uploadedRecords': len(staff_data),
            'errors': errors
        })

    except Exception as e:
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500


def process_excel_image(sheet, row_num, col_index, identifier):
    """Process image from Excel cell and upload to Firebase Storage"""
    try:
        # Get images from the worksheet
        for image in sheet._images:
            # Check if image is in the target cell
            if (image.anchor._from.row + 1 == row_num and 
                image.anchor._from.col == col_index):
                
                # Extract image bytes
                image_bytes = image._data()
                
                if image_bytes:
                    # Create filename using identifier (email or emp_id)
                    safe_identifier = str(identifier).replace('@', '_at_').replace('.', '_dot_')
                    filename = f'staff_photos/{safe_identifier}_photo.jpg'
                    
                    # Upload to Firebase Storage
                    blob = bucket.blob(filename)
                    blob.upload_from_string(image_bytes, content_type='image/jpeg')
                    
                    # Make public and get URL
                    blob.make_public()
                    return blob.public_url
        
        return ''
    except Exception as e:
        print(f'Error processing image: {e}')
        return ''


# Progress endpoint for real-time updates during upload
@app.route('/api/upload_progress/<task_id>', methods=['GET'])
@admin_required
def get_upload_progress(task_id):
    # This would be used with a task queue like Celery for real progress tracking
    # For now, return a simple response
    return jsonify({
        'status': 'processing',
        'progress': 50,
        'message': 'Processing records...'
    })


# Endpoint used by web client to create firebase auth user if login fails (mirrors Flutter fallback)
@app.route('/api/create_if_staff', methods=['POST'])
def create_if_staff():
    data = request.json or {}
    email = (data.get('email') or '').strip()
    password = (data.get('password') or '').strip()
    if not email or not password:
        return jsonify({'error': 'email and password required'}), 400

    # Find staff doc with this email
    try:
        q = db.collection('staff').where('email', '==', email).limit(1).get()
        if not q:
            return jsonify({'found': False, 'message': 'No staff record found'}), 404
        staff_doc = q[0]
        staff = staff_doc.to_dict()
        # Accept phone or mobileNo or Phone
        stored_phone = str(staff.get('mobileNo') or staff.get('Phone') or staff.get('phone') or '')
        if stored_phone == password:
            # create Firebase Auth user
            try:
                new_user = auth.create_user(email=email, password=password)
                db.collection('users').document(new_user.uid).set({
                    'email': email,
                    'isAdmin': False,
                    'staffId': staff.get('slNo') or staff.get('SlNo') or '',
                    'createdAt': firestore.SERVER_TIMESTAMP
                })
                return jsonify({'created': True, 'uid': new_user.uid}), 201
            except Exception as e:
                return jsonify({'error': 'Could not create user', 'detail': str(e)}), 500
        else:
            return jsonify({'found': True, 'match': False, 'message': 'Password did not match staff phone'}), 403
    except Exception as e:
        return jsonify({'error': 'Server error', 'detail': str(e)}), 500

# Protected endpoint to list staff entries (for web UI)
@app.route('/api/staffs', methods=['GET'])
@login_required
def list_staffs():
    try:
        # Debug: print the Authorization header to see if token is sent
        from flask import request
        print("Authorization header received:", request.headers.get("Authorization"))

        docs = db.collection('staff').stream()
        staff_list = []
        for d in docs:
            data = d.to_dict()
            data['id'] = d.id
            staff_list.append(data)

        return jsonify({'staffs': staff_list})
    except Exception as e:
        return jsonify({'error': 'Failed to fetch staffs', 'detail': str(e)}), 500


# Use web_login_required instead of login_required for HTML pages
@app.route('/staff_list.html')
@web_login_required
def staff_list_page_html():
    return render_template('staff_list.html')

@app.route('/admin.html')
@web_login_required
def admin_page_html():
    return render_template('admin.html')


# Browser file upload endpoint — processes files server-side
@app.route('/api/upload_file', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    if not allowed_file(f.filename):
        return jsonify({'error': 'File type not allowed'}), 400

    filename = secure_filename(f.filename)
    local_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(local_path)

    # Optional: run ported processing logic
    try:
        result = process_uploaded_file(local_path)  # implement in processing/file_processors.py
    except Exception as e:
        result = {'error': 'processing_failed', 'detail': str(e)}

    # upload to firebase storage
    blob = bucket.blob(f'uploads/{filename}')
    blob.upload_from_filename(local_path)
    # optionally set public or generate signed url — be cautious with privacy
    try:
        blob.make_public()
        storage_url = blob.public_url
    except Exception:
        storage_url = None

    return jsonify({'filename': filename, 'storage_url': storage_url, 'process_result': result})

# Serve static files (if needed)
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)


# Add this new endpoint to your existing app.py after the other API routes

@app.route('/api/staff/<staff_id>', methods=['DELETE'])
@admin_required
def delete_staff(staff_id):
    """
    Delete a staff member (admin only)
    Also optionally delete the associated Firebase Auth user
    """
    try:
        # First, get the staff document to get email for user deletion
        staff_doc = db.collection('staff').document(staff_id).get()
        if not staff_doc.exists:
            return jsonify({'error': 'Staff member not found'}), 404
            
        staff_data = staff_doc.to_dict()
        staff_email = staff_data.get('email')
        
        # Delete the staff document
        db.collection('staff').document(staff_id).delete()
        
        # Optionally delete the associated Firebase Auth user and users collection document
        deleted_auth_user = False
        if staff_email:
            try:
                # Find user by email
                user_record = auth.get_user_by_email(staff_email)
                
                # Delete from users collection
                db.collection('users').document(user_record.uid).delete()
                
                # Delete Firebase Auth user
                auth.delete_user(user_record.uid)
                deleted_auth_user = True
                
            except auth.UserNotFoundError:
                # User doesn't exist in Firebase Auth, that's okay
                pass
            except Exception as user_delete_error:
                # Log error but don't fail the staff deletion
                print(f"Failed to delete auth user for {staff_email}: {user_delete_error}")
        
        return jsonify({
            'success': True, 
            'message': 'Staff member deleted successfully',
            'deleted_auth_user': deleted_auth_user,
            'staff_id': staff_id
        })
        
    except Exception as e:
        return jsonify({'error': 'Failed to delete staff member', 'detail': str(e)}), 500


# Add this endpoint to your app.py after the other staff routes

@app.route('/api/staff/<staff_id>/type', methods=['PUT'])
@admin_required
def update_staff_type(staff_id):
    """
    Update staff type (Teaching Staff, Non Teaching Staff, or Retired)
    """
    try:
        data = request.get_json()
        new_type = data.get('type')
        
        if not new_type:
            return jsonify({'error': 'Staff type is required'}), 400
        
        # Validate the type
        valid_types = ['Teaching Staff', 'Non Teaching Staff', 'Retired']
        if new_type not in valid_types:
            return jsonify({'error': f'Invalid staff type. Must be one of: {", ".join(valid_types)}'}), 400
        
        # Get the staff document
        staff_doc = db.collection('staff').document(staff_id).get()
        if not staff_doc.exists:
            return jsonify({'error': 'Staff member not found'}), 404
        
        # Update the staff type
        db.collection('staff').document(staff_id).update({
            'type': new_type,
            'updatedAt': firestore.SERVER_TIMESTAMP
        })
        
        return jsonify({
            'success': True,
            'message': f'Staff type updated to {new_type}',
            'staff_id': staff_id,
            'new_type': new_type
        })
        
    except Exception as e:
        print(f"Error updating staff type: {e}")
        return jsonify({'error': 'Failed to update staff type', 'detail': str(e)}), 500


@app.route('/api/staff/bulk_delete', methods=['POST'])
@admin_required
def bulk_delete_staff():
    """
    Delete multiple staff members (admin only)
    Expects JSON: {"staff_ids": ["id1", "id2", "id3"]}
    """
    try:
        data = request.json or {}
        staff_ids = data.get('staff_ids', [])
        
        if not staff_ids or not isinstance(staff_ids, list):
            return jsonify({'error': 'staff_ids array required'}), 400
            
        deleted_count = 0
        deleted_auth_users = 0
        errors = []
        
        for staff_id in staff_ids:
            try:
                # Get staff document
                staff_doc = db.collection('staff').document(staff_id).get()
                if not staff_doc.exists:
                    errors.append({'staff_id': staff_id, 'error': 'Staff not found'})
                    continue
                    
                staff_data = staff_doc.to_dict()
                staff_email = staff_data.get('email')
                
                # Delete staff document
                db.collection('staff').document(staff_id).delete()
                deleted_count += 1
                
                # Try to delete associated auth user
                if staff_email:
                    try:
                        user_record = auth.get_user_by_email(staff_email)
                        db.collection('users').document(user_record.uid).delete()
                        auth.delete_user(user_record.uid)
                        deleted_auth_users += 1
                    except auth.UserNotFoundError:
                        pass
                    except Exception as user_error:
                        errors.append({'staff_id': staff_id, 'error': f'Auth deletion failed: {user_error}'})
                        
            except Exception as e:
                errors.append({'staff_id': staff_id, 'error': str(e)})
        
        return jsonify({
            'success': True,
            'deleted_staff': deleted_count,
            'deleted_auth_users': deleted_auth_users,
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'error': 'Bulk delete failed', 'detail': str(e)}), 500

@app.route('/api/staff/test_admin_check', methods=['GET'])
@admin_required
def test_admin_check():
    """
    Simple endpoint to test if user has admin privileges
    Used by frontend to determine if admin features should be shown
    """
    return jsonify({'isAdmin': True})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))